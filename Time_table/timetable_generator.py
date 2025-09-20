#!/usr/bin/env python3
# -- coding: utf-8 --
"""
timetable_full_locking_fixed.py

Added feature:
 - When the user enters semester (3/5/7) while creating division entries,
   the script will prompt to optionally mark a weekly whole-day free for that
   sem/div (allowed options: sem 3/5 -> Mon or Sat, sem 7 -> Fri, Sat or both).
 - Free days are applied to the division timetable before scheduling so no
   classes will be placed on those days. The free-day label is:
     "COMPETITIVE EXAM/SUNCLUBS/SPORT (Sem{sem} Div{div})"

Other behaviors, locking logic and exports are left unchanged.
"""
from __future__ import annotations
import logging
import random
import re
import sys
import pandas as pd
from collections import OrderedDict
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

# ---------- Logging ----------
LOG_FILE = "logs.txt"
logger = logging.getLogger("timetable_full_locking_fixed")
logger.setLevel(logging.DEBUG)

# File handler
fh = logging.FileHandler(LOG_FILE, mode="w", encoding="utf-8")
fh.setLevel(logging.DEBUG)
fh_formatter = logging.Formatter('%(asctime)s | %(levelname)-7s | %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
fh.setFormatter(fh_formatter)
logger.addHandler(fh)

# Console handler (info)
ch = logging.StreamHandler(sys.stdout)
ch.setLevel(logging.INFO)
ch_formatter = logging.Formatter('[%(levelname)s] %(message)s')
ch.setFormatter(ch_formatter)
logger.addHandler(ch)

logger.info("Timetable generator started")

# ---------- Config (respect user timings exactly) ----------
DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]

SHIFT_SLOTS = {
    "8-3": [
        "8-8:45","8:45-9:45",
        "9:45-10:00 Short Break",
        "10:00-11:00","11:00-12:00",
        "12:00-12:45 Lunch Break",
        "12:45-1:45","1:45-2:45"
    ],
    "10-5": [
        "10:00-11:00","11:00-12:00",
        "12:00-12:45 Lunch Break",
        "12:45-1:45","1:45-2:45",
        "2:45-3:00 Short Break",
        "3-4","4-5"
    ]
}

DESIGNATION_MENU = {"1":"Professor","2":"Assistant Professor","3":"Jr Assistant Professor"}
SHIFT_MENU       = {"1":"8-3","2":"10-5"}
TEACHING_TYPE_MENU={"1":"Lab","2":"Theory","3":"Both"}

# Global registries
FACULTY_SUBJECT_COURSE = {}   # (faculty_short, sem, div, subject, type) -> course_code
FACULTY_FULLNAME = {}        # short -> full name

# Free day config & storage
FREE_DAY_LABEL = "COMPETITIVE EXAM/SUNCLUBS/SPORT"
# ---------- Slot canonicalization utilities ----------
_RE_TIME_PART = re.compile(r'^(\d{1,2})(?::(\d{1,2}))?$')

def parse_time_token(tok: str, shift_hint: str=None) -> int:
    m = _RE_TIME_PART.match(tok.strip())
    if not m:
        raise ValueError(f"Cannot parse time token: {tok!r}")
    hh = int(m.group(1))
    mm = int(m.group(2) or 0)
    # heuristics:
    if hh < 8:
        hh += 12
    return hh * 60 + mm

def slot_label_to_canonical(slot_label: str, shift: str = None) -> tuple[int,int]:
    main = slot_label.split()[0]
    parts = main.split('-')
    if len(parts) != 2:
        raise ValueError(f"Unexpected slot format: {slot_label!r}")
    a, b = parts[0].strip(), parts[1].strip()
    start = parse_time_token(a)
    end = parse_time_token(b)
    return (start, end)

def canonical_to_string(pair: tuple[int,int]) -> str:
    s,e = pair
    def fmt(m):
        hh = m // 60
        mm = m % 60
        return f"{hh:02d}:{mm:02d}"
    return f"{fmt(s)}-{fmt(e)}"

CANONICAL_MAP = {}
REVERSE_CANONICAL_MAP = {}
for sh, slots in SHIFT_SLOTS.items():
    cmap = {}
    rmap = {}
    for sl in slots:
        try:
            canon = slot_label_to_canonical(sl, sh)
        except Exception:
            canon = None
        cmap[sl] = canon
        if canon is not None:
            rmap.setdefault(canon, []).append(sl)
    CANONICAL_MAP[sh] = cmap
    REVERSE_CANONICAL_MAP[sh] = rmap

def slots_equivalent(shift_a, label_a, shift_b, label_b) -> bool:
    ca = CANONICAL_MAP[shift_a].get(label_a)
    cb = CANONICAL_MAP[shift_b].get(label_b)
    return (ca is not None) and (cb is not None) and (ca[0] == cb[0] and ca[1] == cb[1])

def pair_slots_equivalent(shift_a, a1, a2, shift_b, b1, b2) -> bool:
    return slots_equivalent(shift_a, a1, shift_b, b1) and slots_equivalent(shift_a, a2, shift_b, b2)

# ---------- Timetable table helpers ----------
def empty_table_for_shift(shift):
    table = {d: {} for d in DAYS}
    for d in DAYS:
        for s in SHIFT_SLOTS[shift]:
            table[d][s] = s if ("Break" in s or "Lunch" in s) else ""
    return table

def consecutive_pairs_for_shift(shift):
    slots = [s for s in SHIFT_SLOTS[shift]]
    pairs = []
    for i in range(len(slots)-1):
        a, b = slots[i], slots[i+1]
        if ("Break" in a or "Lunch" in a or "Break" in b or "Lunch" in b):
            continue
        if CANONICAL_MAP[shift].get(a) is None or CANONICAL_MAP[shift].get(b) is None:
            continue
        pairs.append((a,b))
    return pairs

def free_slot(tbl, day, slot):
    # True only when the cell is an empty string (available).
    val = tbl[day].get(slot, "")
    if not isinstance(val, str):
        return False
    v = val.strip()
    if v == "":
        return True
    up = v.upper()
    # detect holiday / competitive / sports markers
    if FREE_DAY_LABEL.upper() in up or "COMPETITIVE" in up or "SUNCLUBS" in up or "SPORT" in up or "HOLIDAY" in up:
        return False
    # block breaks, lunches, merges
    if "BREAK" in up or "LUNCH" in up or v == "MERGE":
        return False
    # any other non-empty text means occupied
    return False


def free_pair(tbl, day, s1, s2):
    return free_slot(tbl, day, s1) and free_slot(tbl, day, s2)

def extract_subject_from_cell(text):
    if not isinstance(text, str) or not text.strip():
        return None
    t = text.strip()
    t = re.sub(r'\[[^\]]+\]$', '', t).strip()
    m = re.match(r'^\s*([^\(]+?)\s+(?:Lab|\()', t)
    if m:
        return m.group(1).strip()
    m2 = re.match(r'^\s*([^\(]+?)\s*(?:\(|-)', t)
    if m2:
        return m2.group(1).strip()
    return t.split()[0] if t.split() else t

def day_has_subject(tbl, day, subject):
    target = (subject or "").strip().lower()
    if not target:
        return False
    for v in tbl[day].values():
        if not isinstance(v, str): continue
        if target in v.lower():
            return True
    return False

def day_has_division(tbl, day, sem, div):
    needle_sem = f"sem{sem}".lower()
    needle_div = f"div{div}".lower()
    for v in tbl[day].values():
        if not isinstance(v, str): continue
        low = v.lower().replace(" ", "")
        if needle_sem in low and needle_div in low:
            return True
    return False

# ---------- shift compatibility rules ----------
def is_slot_allowed_for_10_5_on_8_3(slot_label: str) -> bool:
    canon = None
    for sh in CANONICAL_MAP:
        if slot_label in CANONICAL_MAP[sh]:
            canon = CANONICAL_MAP[sh][slot_label]
            break
    if canon is None:
        return False
    start, end = canon
    return start >= 600

def division_slot_allowed_for_faculty(fac_shift: str, div_shift: str, slot_label: str) -> bool:
    if fac_shift == "10-5" and div_shift == "8-3":
        return is_slot_allowed_for_10_5_on_8_3(slot_label)
    return True

def division_pair_allowed_for_faculty(fac_shift: str, div_shift: str, pair: tuple) -> bool:
    if fac_shift == "10-5" and div_shift == "8-3":
        return is_slot_allowed_for_10_5_on_8_3(pair[0]) and is_slot_allowed_for_10_5_on_8_3(pair[1])
    return True

# ---------- Input helpers (unchanged semantics) ----------
def input_menu(prompt, menu):
    while True:
        print(prompt)
        for k, v in menu.items():
            print(f"{k}: {v}")
        ch = input("Enter choice: ").strip()
        if ch in menu:
            return menu[ch]
        print("Invalid. Try again.")

def input_int(prompt, min_val=None, max_val=None):
    while True:
        try:
            v = int(input(prompt).strip())
            if min_val is not None and v < min_val:
                print("Out of range.")
                continue
            if max_val is not None and v > max_val:
                print("Out of range.")
                continue
            return v
        except Exception:
            print("Enter valid integer.")

def normalize_token(s: str) -> str:
    return re.sub(r'\s+','', (s or "").strip()).upper()

def parse_batches_input(raw: str):
    raw = (raw or "").strip()
    if raw == "":
        return ["B1"], False
    if "/" in raw and "," not in raw:
        return [raw.replace(" ", "")], True
    parts = [normalize_token(x) for x in re.split(r'[ ,]+', raw) if x.strip()]
    if not parts: parts = ["B1"]
    return parts, False

# ---------- Data entry ----------
# Global store for holidays
# Days map for selection
# Days map for selection
DAYS_MAP = {
    "1": "Mon",
    "2": "Tue",
    "3": "Wed",
    "4": "Thu",
    "5": "Fri",
    "6": "Sat"
}

# Stores mapping (sem,div) -> list_of_days_selected
FREE_DAY_SETTINGS: dict[tuple[str,str], list[str]] = {}

def ask_and_record_free_days_for_division(sem: str, div: str):
    """
    Prompt user to mark free/holiday days for Sem/Div.
    Updates FREE_DAY_SETTINGS[(sem,div)] = [days...]
    """

    # ask how many holidays for this division (directly)
    while True:
        try:
            num_holidays = int(input(f"\nEnter number of holidays for Sem {sem} Div {div} (0-6): ").strip())
            if 0 <= num_holidays <= len(DAYS_MAP):
                break
            else:
                print("⚠️ Please enter a number between 0 and 6.")
        except ValueError:
            print("⚠️ Enter a valid number.")

    holidays_list = []
    if num_holidays > 0:
        print("Select holiday days (Mon–Sat):")
        for k, v in DAYS_MAP.items():
            print(f"{k}: {v}")

        for i in range(num_holidays):
            while True:
                hday = input(f"Enter holiday {i+1} for Sem {sem} Div {div} (1-6): ").strip()
                if hday in DAYS_MAP and DAYS_MAP[hday] not in holidays_list:
                    holidays_list.append(DAYS_MAP[hday])
                    break
                else:
                    print("⚠️ Invalid or duplicate choice, try again.")

    # Save to settings
    FREE_DAY_SETTINGS[(str(sem), normalize_token(div))] = holidays_list
    logger.info("Recorded holidays for Sem%s Div%s => %s", sem, div, holidays_list)



def get_faculty_data():
    """
    Interactive input compatible with prior script.
    Additionally asks per-division free-day choices for sem 3/5/7.
    """
    print("Timetable generator starting. Enter inputs below.")
    faculty_list = []
    n = input_int("Enter number of faculty members: ", 1)
    for i in range(n):
        print(f"\n--- Enter details for Faculty {i+1} ---")
        name = input("Faculty short name (used for filenames, e.g. 'MSK'): ").strip()
        full_name = input("Faculty Full Name (for reports): ").strip()
        FACULTY_FULLNAME[name] = full_name or name
        designation = input_menu("Choose Designation:", DESIGNATION_MENU)
        shift = input_menu("Choose Faculty Shift:", SHIFT_MENU)
        weekly_hours = input_int("Weekly Teaching Hours: ", 1)

        subjects = []
        while True:
            sem = input("Semester (3 / 5 / 7 or 'done'): ").strip()
            if sem.lower() == "done": break
            if sem not in {"3","5","7"}:
                print("Enter 3 / 5 / 7 or 'done'."); continue

            sub_type = input_menu("Select Type:", TEACHING_TYPE_MENU)
            if sub_type in ("Lab", "Both"):
                divs_raw = input("Lab Divisions (comma-separated, e.g., A,B): ")
                divs = [normalize_token(x) for x in re.split(r'[ ,]+', divs_raw) if x.strip()]
                for div in divs:
                    div_shift = input_menu(f"Choose Shift for Division {div}:", SHIFT_MENU)
                    # Ask free-day setting for this sem/div (NEW)
                    ask_and_record_free_days_for_division(sem, div)

                    num_labs = input_int(f"How many labs per week for Division {div}? (Each = 2 hrs): ", 1)
                    lab_subject = input(f"Lab Subject for Division {div}: ").strip()
                    course_code = input(f"Course Code for '{lab_subject}' (Division {div}, Semester {sem}): ").strip()
                    batches_raw = input(f"Lab Batches for Division {div} (comma-separated or '/' to group): ").strip()
                    batches, grouped = parse_batches_input(batches_raw)
                    subjects.append({
                        "Type": "Lab",
                        "Semester": sem,
                        "Division": div,
                        "Div_Shift": div_shift,
                        "Subject": lab_subject,
                        "Course_Code": course_code,
                        "Num_Labs": num_labs,
                        "Batches": batches,
                        "Batches_Grouped": grouped,
                        "Placed": 0
                    })
                    FACULTY_SUBJECT_COURSE[(name, sem, div, lab_subject, "Lab")] = course_code

            if sub_type in ("Theory", "Both"):
                divs_raw = input("Theory Divisions (comma-separated, e.g., A,B): ")
                divs = [normalize_token(x) for x in re.split(r'[ ,]+', divs_raw) if x.strip()]
                for div in divs:
                    div_shift = input_menu(f"Choose Shift for Division {div}:", SHIFT_MENU)
                    # Ask free-day setting for this sem/div (NEW)
                    ask_and_record_free_days_for_division(sem, div)

                    subject = input(f"Subject Name for Division {div}: ").strip()
                    course_code = input(f"Course Code for '{subject}' (Division {div}, Semester {sem}): ").strip()
                    theory_classes = input_int(f"No. of Theory Classes per week for Division {div}: ", 1)
                    subjects.append({
                        "Type": "Theory",
                        "Semester": sem,
                        "Division": div,
                        "Div_Shift": div_shift,
                        "Subject": subject,
                        "Course_Code": course_code,
                        "Theory_Classes": theory_classes,
                        "Placed": 0
                    })
                    FACULTY_SUBJECT_COURSE[(name, sem, div, subject, "Theory")] = course_code

        faculty_list.append({
            "Name": name,
            "Full_Name": FACULTY_FULLNAME.get(name, name),
            "Designation": designation,
            "Shift": shift,
            "Weekly_Hours": weekly_hours,
            "Subjects": subjects
        })
    return faculty_list

# ---------- Low-level placement (accurate canonical matching) ----------
def ensure_div_table(dtables, sem, div, div_shift):
    key = (sem, div)
    if key not in dtables:
        tbl = empty_table_for_shift(div_shift)

        # Apply division holidays from FREE_DAY_SETTINGS
        holidays = FREE_DAY_SETTINGS.get((str(sem), normalize_token(div)), [])
        for hday in holidays:
            for slot in SHIFT_SLOTS[div_shift]:
                if "Break" in slot or "Lunch" in slot:
                    continue
                tbl[hday][slot] = f"{FREE_DAY_LABEL} (Sem{sem} Div{div})"

        dtables[key] = {"shift": div_shift, "table": tbl}

    return dtables[key]["table"], dtables[key]["shift"]

# ---------- Holiday check helper ----------
def is_division_holiday(sem, div, day):
    """
    Normalize sem/div keys and check if this day is marked as a holiday
    in FREE_DAY_SETTINGS.
    """
    key = (str(sem), normalize_token(div))
    return day in FREE_DAY_SETTINGS.get(key, [])


def lock_theory(ftbl, dtbl, fshift, dshift, fname, sem, div, subject, avoid_dup=True):
    for day in random.sample(DAYS, len(DAYS)):
        if avoid_dup and day_has_division(dtbl, day, sem, div):
            continue
        if avoid_dup and day_has_subject(dtbl, day, subject):
            continue
        # skip if division day is a holiday
        if is_division_holiday(sem, div, day):
            logger.info("[HOLIDAY-SKIP] Skipping Sem%s Div%s on %s", sem, div, day)
            continue


        for fslot in SHIFT_SLOTS[fshift]:
            if "Break" in fslot or "Lunch" in fslot: continue
            if not free_slot(ftbl, day, fslot): continue
            for dslot in SHIFT_SLOTS[dshift]:
                if "Break" in dslot or "Lunch" in dslot: continue
                if not free_slot(dtbl, day, dslot): continue
                if not slots_equivalent(fshift, fslot, dshift, dslot): continue
                if not division_slot_allowed_for_faculty(fshift, dshift, dslot): continue
                ftbl[day][fslot] = f"{subject} (Sem{sem} Div{div})"
                dtbl[day][dslot] = f"{subject} ({fname})"
                logger.info(f"[SUCCESS] Theory: {subject} assigned by {fname} -> Sem{sem} Div{div} at {day} F={fslot} D={dslot}")
                return True
    for day in random.sample(DAYS, len(DAYS)):
        for fslot in SHIFT_SLOTS[fshift]:
            if "Break" in fslot or "Lunch" in fslot: continue
            if not free_slot(ftbl, day, fslot): continue
            for dslot in SHIFT_SLOTS[dshift]:
                if "Break" in dslot or "Lunch" in dslot: continue
                if not free_slot(dtbl, day, dslot): continue
                if not slots_equivalent(fshift, fslot, dshift, dslot): continue
                if not division_slot_allowed_for_faculty(fshift, dshift, dslot): continue
                ftbl[day][fslot] = f"{subject} (Sem{sem} Div{div})"
                dtbl[day][dslot] = f"{subject} ({fname})"
                logger.info(f"[SUCCESS] Theory (fallback): {subject} assigned by {fname} -> Sem{sem} Div{div} at {day} F={fslot} D={dslot}")
                return True
    logger.debug(f"[TRY-FAIL] Theory: {subject} not placed (yet) for {fname} Sem{sem} Div{div}")
    return False

def lock_lab(ftbl, dtbl, fshift, dshift, fname, sem, div, subject, batch_label, avoid_dup=True):
    fpairs = consecutive_pairs_for_shift(fshift)
    dpairs = consecutive_pairs_for_shift(dshift)
    for day in random.sample(DAYS, len(DAYS)):
        if avoid_dup and day_has_division(dtbl, day, sem, div):
            continue
        if avoid_dup and day_has_subject(dtbl, day, subject):
            continue
        # skip if division day is a holiday
        if is_division_holiday(sem, div, day):
            logger.info("[HOLIDAY-SKIP] Skipping Sem%s Div%s on %s", sem, div, day)
            continue


        for (fs1, fs2) in fpairs:
            if not free_pair(ftbl, day, fs1, fs2): continue
            for (ds1, ds2) in dpairs:
                if not free_pair(dtbl, day, ds1, ds2): continue
                if not pair_slots_equivalent(fshift, fs1, fs2, dshift, ds1, ds2): continue
                if not division_pair_allowed_for_faculty(fshift, dshift, (ds1, ds2)): continue
                ftbl[day][fs1] = f"{subject} Lab (Sem{sem} Div{div}) [{batch_label}]"; ftbl[day][fs2] = "MERGE"
                dtbl[day][ds1] = f"{subject} Lab ({fname}) [{batch_label}]"; dtbl[day][ds2] = "MERGE"
                logger.info(f"[SUCCESS] Lab: {subject} ({batch_label}) assigned by {fname} -> Sem{sem} Div{div} at {day} F=({fs1},{fs2}) D=({ds1},{ds2})")
                return True
    for fday in random.sample(DAYS, len(DAYS)):
        for (fs1, fs2) in fpairs:
            if not free_pair(ftbl, fday, fs1, fs2): continue
            for dday in random.sample(DAYS, len(DAYS)):
                if avoid_dup and day_has_division(dtbl, dday, sem, div):
                    continue
                for (ds1, ds2) in dpairs:
                    if not free_pair(dtbl, dday, ds1, ds2): continue
                    if not pair_slots_equivalent(fshift, fs1, fs2, dshift, ds1, ds2): continue
                    if not division_pair_allowed_for_faculty(fshift, dshift, (ds1, ds2)): continue
                    ftbl[fday][fs1] = f"{subject} Lab (Sem{sem} Div{div}) [{batch_label}]"; ftbl[fday][fs2] = "MERGE"
                    dtbl[dday][ds1] = f"{subject} Lab ({fname}) [{batch_label}]"; dtbl[dday][ds2] = "MERGE"
                    logger.info(f"[SUCCESS-FLEX] Lab: {subject} ({batch_label}) assigned by {fname} -> Sem{sem} Div{div} Fday={fday} Dday={dday}")
                    return True
    logger.debug(f"[TRY-FAIL] Lab: {subject} not placed (yet) for {fname} Sem{sem} Div{div} [{batch_label}]")
    return False

# ---------- Force (deterministic) passes (use canonical equality) ----------
def force_place_theory(ftbl, dtbl, fshift, dshift, fname, sem, div, subject):
    logger_force = logging.getLogger("force_theory")
    fslots = [s for s in SHIFT_SLOTS[fshift] if "Break" not in s and "Lunch" not in s]
    dslots = [s for s in SHIFT_SLOTS[dshift] if "Break" not in s and "Lunch" not in s]

    for day in DAYS:
        # ⬇️ check here first
        if is_division_holiday(sem, div, day):
            logger.info("[HOLIDAY-SKIP] Skipping Sem%s Div%s on %s", sem, div, day)
            continue

        for fs in fslots:
            if not free_slot(ftbl, day, fs): 
                continue
            for ds in dslots:
                if not free_slot(dtbl, day, ds): continue
                if not slots_equivalent(fshift, fs, dshift, ds): continue
                if not division_slot_allowed_for_faculty(fshift, dshift, ds): continue
                ftbl[day][fs] = f"{subject} (Sem{sem} Div{div})"
                dtbl[day][ds] = f"{subject} ({fname})"
                logger_force.warning("[FORCE] THEORY forced: %s -> Sem%s Div%s at %s F=%s D=%s",
                                    fname, sem, div, day, fs, ds)
                return True


    for dday in DAYS:
        for ds in dslots:
            if not free_slot(dtbl, dday, ds): continue
            for fday in DAYS:
                for fs in fslots:
                    if not free_slot(ftbl, fday, fs): continue
                    if not slots_equivalent(fshift, fs, dshift, ds): continue
                    if not division_slot_allowed_for_faculty(fshift, dshift, ds): continue
                    ftbl[fday][fs] = f"{subject} (Sem{sem} Div{div})"; dtbl[dday][ds] = f"{subject} ({fname})"
                    logger_force.warning("[FORCE-RELAX] THEORY forced (relaxed): %s -> Sem%s Div%s Fday=%s Dday=%s", fname, sem, div, fday, dday)
                    return True

    logger_force.error("[FAILED FORCE] THEORY unable to force-place %s Sem%s Div%s", subject, sem, div)
    return False

def force_place_lab(ftbl, dtbl, fshift, dshift, fname, sem, div, subject, batch_label):
    logger_force = logging.getLogger("force_lab")
    fpairs = consecutive_pairs_for_shift(fshift)
    dpairs = consecutive_pairs_for_shift(dshift)

    for day in DAYS:
        # ⬇️ holiday check here
        if is_division_holiday(sem, div, day):
            logger.info("[HOLIDAY-SKIP] Skipping Sem%s Div%s on %s", sem, div, day)
            continue

        for (fs1, fs2) in fpairs:
            if not free_pair(ftbl, day, fs1, fs2): continue
            for (ds1, ds2) in dpairs:
                if not free_pair(dtbl, day, ds1, ds2): continue
                if not pair_slots_equivalent(fshift, fs1, fs2, dshift, ds1, ds2): continue
                if not division_pair_allowed_for_faculty(fshift, dshift, (ds1, ds2)): continue
                ftbl[day][fs1] = f"{subject} Lab (Sem{sem} Div{div}) [{batch_label}]"
                ftbl[day][fs2] = "MERGE"
                dtbl[day][ds1] = f"{subject} Lab ({fname}) [{batch_label}]"
                dtbl[day][ds2] = "MERGE"
                logger_force.warning("[FORCE] LAB forced: %s -> Sem%s Div%s at %s F=(%s,%s) D=(%s,%s)",
                                    fname, sem, div, day, fs1, fs2, ds1, ds2)
                return True

                

    for dday in DAYS:
        for (ds1, ds2) in dpairs:
            if not free_pair(dtbl, dday, ds1, ds2): continue
            for fday in DAYS:
                for (fs1, fs2) in fpairs:
                    if not free_pair(ftbl, fday, fs1, fs2): continue
                    if not pair_slots_equivalent(fshift, fs1, fs2, dshift, ds1, ds2): continue
                    ftbl[fday][fs1] = f"{subject} Lab (Sem{sem} Div{div}) [{batch_label}]"; ftbl[fday][fs2] = "MERGE"
                    dtbl[dday][ds1] = f"{subject} Lab ({fname}) [{batch_label}]"; dtbl[dday][ds2] = "MERGE"
                    logger_force.warning("[FORCE-RELAX] LAB forced (relaxed): %s -> Sem%s Div%s Fday=%s Dday=%s", fname, sem, div, fday, dday)
                    return True

    logger_force.error("[FAILED FORCE] LAB unable to place %s Sem%s Div%s", subject, sem, div)
    return False
def assign_subjects_for_faculty(f, ftables, dtables):
    fname = f["Name"]
    fshift = f["Shift"]
    fsubjects = f["Subjects"]

    if fname not in ftables:
        ftables[fname] = empty_table_for_shift(fshift)
    ftbl = ftables[fname]

    pending = []

    # Labs
    for entry in [s for s in fsubjects if s["Type"] == "Lab"]:
        sem = entry["Semester"]
        div = entry["Division"]
        sub = entry["Subject"]
        dshift = entry["Div_Shift"]
        dtbl, _ = ensure_div_table(dtables, sem, div, dshift)
        batches = entry["Batches"] if not entry.get("Batches_Grouped", False) else [entry["Batches"][0]]
        for batch in batches:
            for _ in range(entry["Num_Labs"]):
                ok = lock_lab(ftbl, dtbl, fshift, dshift, fname, sem, div, sub, batch, avoid_dup=True)
                if not ok:
                    pending.append({
                        "Type":"Lab","Semester":sem,"Division":div,"Div_Shift":dshift,
                        "Subject":sub,"Batch":batch,"Faculty":fname,"FShift":fshift
                    })

    # Theories
    for entry in [s for s in fsubjects if s["Type"] == "Theory"]:
        sem = entry["Semester"]
        div = entry["Division"]
        sub = entry["Subject"]
        dshift = entry["Div_Shift"]
        dtbl, _ = ensure_div_table(dtables, sem, div, dshift)
        for _ in range(entry["Theory_Classes"]):
            ok = lock_theory(ftbl, dtbl, fshift, dshift, fname, sem, div, sub, avoid_dup=True)
            if not ok:
                pending.append({
                    "Type":"Theory","Semester":sem,"Division":div,"Div_Shift":dshift,
                    "Subject":sub,"Faculty":fname,"FShift":fshift
                })

    # Force-fill pending tasks
    for task in pending:
        typ = task["Type"]
        sem = task["Semester"]
        div = task["Division"]
        sub = task["Subject"]
        batch = task.get("Batch", None)
        fshift_task = task.get("FShift", fshift)
        dshift = task["Div_Shift"]

        dtbl, _ = ensure_div_table(dtables, sem, div, dshift)

        if typ == "Theory":
            force_place_theory(ftbl, dtbl, fshift_task, dshift, fname, sem, div, sub)
        else:
            force_place_lab(ftbl, dtbl, fshift_task, dshift, fname, sem, div, sub, batch)



# ---------- Apply free-days to division tables (before scheduling) ----------
def apply_free_day_markings_from_inputs(dtables, faculties):
    """
    Walks through faculty inputs to identify divisions and apply FREE_DAY_SETTINGS
    by pre-filling division tables' slots for the selected days.
    """
    # gather unique (sem,div) -> div_shift
    div_shift_map = {}
    for f in faculties:
        for s in f["Subjects"]:
            key = (s["Semester"], s["Division"])
            if key not in div_shift_map:
                div_shift_map[key] = s.get("Div_Shift", "8-3")
    # apply free-day settings
    for (sem, div), days in list(FREE_DAY_SETTINGS.items()):
        key = (sem, div)
        div_shift = div_shift_map.get(key, "8-3")
        dtbl, _ = ensure_div_table(dtables, sem, div, div_shift)
        # fill every slot on each selected day with FREE_DAY_LABEL + sem/div info
        for day in days:
            if day not in dtbl:
                logger.debug("Requested free day %s not in day-list (skipping): %s", day, key)
                continue
            for slot in list(dtbl[day].keys()):
                dtbl[day][slot] = f"{FREE_DAY_LABEL} (Sem{sem} Div{div})"
        logger.info("Applied free-day marking for Sem%s Div%s => %s", sem, div, days)

# ---------- Excel export helpers (same behavior & styling) ----------
def dataframe_from_table(table, shift):
    cols = SHIFT_SLOTS[shift]
    df = pd.DataFrame.from_dict(table, orient="index")
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    df = df[cols]
    df.index.name = "Day/Time"
    return df

COLOR_PALETTE = [
    "FFB3E5FC","FFFFF9C4","FFC8E6C9","FFFFCCBC","FFD7CCC8","FFE1BEE7",
    "FFFFCDD2","FFFFECB3","FFB2EBF2","FFC5CAE9","FFF8BBD0","FFE6EE9C",
    "FFBBDEFB","FFC8E6C9","FFF0F4C3","FFFFF59D","FFB39DDB"
]

def build_subject_color_map(ftables, dtables):
    subjects = OrderedDict()
    def collect_from_table(tbl):
        for day in DAYS:
            row = tbl.get(day, {})
            for v in row.values():
                s = extract_subject_from_cell(v)
                if s and s not in subjects:
                    subjects[s] = None
    for tbl in ftables.values(): collect_from_table(tbl)
    for payload in dtables.values(): collect_from_table(payload["table"])
    mapping = {}
    for i, s in enumerate(list(subjects.keys())):
        mapping[s] = COLOR_PALETTE[i % len(COLOR_PALETTE)]
    return mapping

def save_excel_with_merges_and_summary(
    filename,
    table,
    shift,
    bottom_summary_rows=None,
    bottom_summary_header=None,
    subject_color_map=None,
    header_type=None,
    faculty_obj=None,
    division=None,
    university="",
    school="",
    department="",
    academic=""
):
    safe_table = {}
    for day, row in table.items():
        newrow = {}
        for k, v in row.items():
            if v == "MERGE":
                newrow[k] = ""
            else:
                newrow[k] = v
        safe_table[day] = newrow

    df = dataframe_from_table(safe_table, shift)
    df.to_excel(filename, sheet_name="Timetable", index=True)

    wb = load_workbook(filename)
    ws = wb.active

    if header_type == "faculty" and faculty_obj:
        display_name = faculty_obj.get("Full_Name", faculty_obj["Name"])
        header_lines = [
            (university, 16, True),
            (department, 14, True),
            (f"INDIVIDUAL TIMETABLE ({display_name})", 13, True),
        ]
    elif header_type == "division" and division:
        sem, div, dshift = division
        header_lines = [
            (university, 16, True),
            (department, 14, True),
            (academic, 13, True),
            (f"Sem {sem} – Div {div} – Shift {dshift}", 12, True),
        ]
    else:
        header_lines = []

    for idx, (text, size, bold) in enumerate(header_lines, start=1):
        ws.insert_rows(idx)
        ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=ws.max_column)
        cell = ws.cell(row=idx, column=1)
        cell.value = text
        cell.font = Font(name="Times New Roman", size=size, bold=bold)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 24

    cols = SHIFT_SLOTS[shift]
    header_to_col = {name: idx for idx, name in enumerate(["Day/Time"] + cols, start=1)}

    # merges for lab slots
    for r_idx, day in enumerate(DAYS, start=2 + len(header_lines)):
        row_orig = table.get(day, {})
        for i, col_name in enumerate(cols):
            val = row_orig.get(col_name, "")
            if val == "MERGE":
                left_col_name = cols[i - 1] if i - 1 >= 0 else None
                if left_col_name:
                    left_col = header_to_col[left_col_name]
                    right_col = header_to_col[col_name]
                    left_val = ws.cell(row=r_idx, column=left_col).value
                    if left_val not in (None, ""):
                        try:
                            ws.merge_cells(
                                start_row=r_idx,
                                start_column=left_col,
                                end_row=r_idx,
                                end_column=right_col,
                            )
                        except Exception:
                            pass
                        ws.cell(row=r_idx, column=left_col).alignment = Alignment(
                            horizontal="center", vertical="center", wrap_text=True
                        )

    # coloring timetable cells
    if subject_color_map:
        merged_ranges = list(ws.merged_cells.ranges)

        def fill_for_hex(hexcolor):
            return PatternFill(start_color=hexcolor, end_color=hexcolor, fill_type="solid")

        top_rows_start = 1 + len(header_lines)
        top_rows_end = top_rows_start + len(DAYS) - 1
        for r in range(top_rows_start + 1, top_rows_end + 1):
            for c in range(2, 2 + len(cols)):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                subj = extract_subject_from_cell(val)
                if subj and subj in subject_color_map:
                    fill = fill_for_hex(subject_color_map[subj])
                    applied = False
                    for mr in merged_ranges:
                        if (r >= mr.min_row and r <= mr.max_row) and (
                            c >= mr.min_col and c <= mr.max_col
                        ):
                            for rr in range(mr.min_row, mr.max_row + 1):
                                for cc in range(mr.min_col, mr.max_col + 1):
                                    ws.cell(row=rr, column=cc).fill = fill
                            applied = True
                            break
                    if not applied:
                        cell.fill = fill

    # center align everything
    for row in ws.iter_rows(
        min_row=1, max_row=1 + len(DAYS) + len(header_lines), min_col=1, max_col=1 + len(cols)
    ):
        for c in row:
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # bottom summary
    if bottom_summary_rows:
        start_row = ws.max_row + 2
        if bottom_summary_header:
            for j, h in enumerate(bottom_summary_header, start=1):
                ws.cell(row=start_row, column=j).value = h
                ws.cell(row=start_row, column=j).alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
            start_row += 1
        for rdata in bottom_summary_rows:
            for j, val in enumerate(rdata, start=1):
                ws.cell(row=start_row, column=j).value = val
                ws.cell(row=start_row, column=j).alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
            start_row += 1

    wb.save(filename)
    logger.info("Saved Excel file: %s", filename)

# ---------- Bottom summary builders ----------
def build_faculty_summary_rows(faculty_obj):
    rows = []
    subj_index = {}
    for s in faculty_obj["Subjects"]:
        sem = s["Semester"]; subj = s["Subject"]
        key = (sem, subj)
        if key not in subj_index:
            subj_index[key] = {"Theory": 0, "Labs": 0}
        if s["Type"] == "Theory":
            subj_index[key]["Theory"] += s.get("Theory_Classes", 0)
        elif s["Type"] == "Lab":
            if s.get("Batches_Grouped", False):
                blocks = s.get("Num_Labs", 0)
            else:
                blocks = s.get("Num_Labs", 0) * max(1, len(s.get("Batches", [])))
            subj_index[key]["Labs"] += blocks

    for (sem, subj), counts in subj_index.items():
        theory = counts["Theory"]
        labs = counts["Labs"]
        total = theory + labs
        rows.append([faculty_obj["Name"], faculty_obj.get("Full_Name", faculty_obj["Name"]), sem, subj, theory, labs, total])
    return rows

def build_division_summary_rows(sem, div, dtbl, dtables):
    rows = []
    seen = set()
    for day in DAYS:
        row = dtbl.get(day, {})
        for val in row.values():
            subj = extract_subject_from_cell(val)
            if subj:
                for (fname, fsem, fdiv, fsubject, ftype), code in FACULTY_SUBJECT_COURSE.items():
                    if fsem == sem and fdiv == div and fsubject and fsubject.strip().lower() == subj.strip().lower():
                        fullname = FACULTY_FULLNAME.get(fname, fname)
                        key = (subj, fullname, (code or ""), ftype)
                        if key not in seen:
                            if ftype == "Lab":
                                rows.append([subj + " [Lab]", fullname, code or ""])
                            else:
                                rows.append([subj, fullname, code or ""])
                            seen.add(key)
    return rows

# ---------- Export all ----------
def export_all(ftables, dtables, faculties_input, university="", department="", academic=""):
    subject_color_map = build_subject_color_map(ftables, dtables)

    # faculties
    for f in faculties_input:
        fname = f["Name"]
        tbl = ftables.get(fname, empty_table_for_shift(f["Shift"]))
        fshift = f["Shift"]
        bottom_rows = build_faculty_summary_rows(f)
        bottom_header = ["FacShort","Faculty Full Name","Semester","Subject","Theory Classes","Labs","Total Sessions"]
        filename = f"Faculty_{fname}.xlsx"
        save_excel_with_merges_and_summary(
            filename, tbl, fshift,
            bottom_summary_rows=bottom_rows,
            bottom_summary_header=bottom_header,
            subject_color_map=subject_color_map,
            header_type="faculty",
            faculty_obj=f,
            division=None,
            university=university,
            department=department,
            academic=academic
        )


    # divisions
    for (sem, div), payload in dtables.items():
        dshift = payload["shift"]; tbl = payload["table"]
        bottom_rows = build_division_summary_rows(sem, div, tbl, dtables)
        bottom_header = ["Subject (Lab indicated)","Faculty Full Name","Course Code"]
        filename = f"Sem{sem}_Div{div}.xlsx"
        save_excel_with_merges_and_summary(
            filename, tbl, dshift,
            bottom_summary_rows=bottom_rows,
            bottom_summary_header=bottom_header,
            subject_color_map=subject_color_map,
            header_type="division",
            faculty_obj=None,
            division=(sem, div, dshift),
            university=university,
            department=department,
            academic=academic
        )

#-----main-----#
def main():
    random.seed(7)
    logger.info("Beginning scheduling run at %s", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    # --- get universal header inputs ---
    print("\n--- Enter Universal Header Information ---")
    university = input("University : ").strip()
    department = input("Department Name: ").strip()
    academic   = input("Semester Label / Academic Year (e.g. TIME TABLE – ODD SEMESTER 2025-26): ").strip() or "TIME TABLE – ODD SEMESTER 2025-26"

    faculties = get_faculty_data()
    ftables = {}
    dtables = {}

    # Apply free-day markings BEFORE scheduling so those days are treated as non-free.
    apply_free_day_markings_from_inputs(dtables, faculties)

    for f in faculties:
        assign_subjects_for_faculty(f, ftables, dtables)


    # pass header values to export_all
    export_all(ftables, dtables, faculties,
               university=university, department=department, academic=academic)

    logger.info("Scheduling run complete. Check generated Excel files and %s for logs.", LOG_FILE)
if  __name__ == "__main__":
     main()
















